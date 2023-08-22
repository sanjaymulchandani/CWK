from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import *
from .forms import allWorkshopForm, WorkshopRegistrationForm, ReviewForm, ContactFormCustomized
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from django.http import HttpResponse


# Create your views here.

# Export workshop form data to excel from admin panel

def export_to_excel(modeladmin, request, queryset):
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write the header row
    headers = [field.verbose_name for field in modeladmin.model._meta.fields]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Write the data rows
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(modeladmin.model._meta.fields, 1):
            col_letter = get_column_letter(col_num)
            value = getattr(obj, field.name)
            ws[f"{col_letter}{row_num}"] = str(value)  # Convert value to string

    # Save the workbook to a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=workshop_records.xlsx'
    wb.save(response)

    return response


def home_view(request):
    hero_home = heroBanner.objects.all()
    trending_workshops = allWorkshop.objects.filter(trending=True)
    reviews = Review.objects.filter(display=True).order_by('?')[:3]
    ad_banner = AdBanner.objects.all()
    promo_banner = PromoBanner.objects.all()
    return render(request, 'home_view.html', context={"hero_home": hero_home,
                                                      "trending_workshops": trending_workshops,
                                                      "reviews": reviews,
                                                      "ad_banner": ad_banner,
                                                      "promo_banner": promo_banner,})


def all_workshop(request):
    all_workshop = allWorkshop.objects.order_by('-date_created')
    return render(request, 'all_workshops.html', context={"all_workshop": all_workshop})


def contact_view(request):
    contact_content = ContactContent.objects.all()

    if request.method == 'POST':
        form = ContactFormCustomized(request.POST)
        if form.is_valid():
            form.save(commit=False)
            messages.success(request, 'Form submitted!')
    else:
        form = ContactFormCustomized()

    context = {
        'contact_content': contact_content,
        'form': form
    }
    return render(request, 'contact_view.html', context)

def about_view(request):
    about_content = About.objects.all()
    context = {"about_content": about_content}
    return render(request, 'about_view.html', context)


def business_view(request):
    return render(request, 'business_view.html')


def blog_view(request):
    return render(request, 'blog_view.html')


def header_links(request, workshop_id):
    workshop = get_object_or_404(allWorkshop, id=workshop_id)
    return render(request, 'includes\header.html', {'workshop': workshop})


def create_workshop(request):
    if request.method == 'POST':
        form = allWorkshopForm(request.POST, request.FILES)
        if form.is_valid():
            workshop = form.save()
            return redirect('workshop_detail', pk=workshop.pk)
    else:
        form = allWorkshopForm()
    return render(request, 'create_workshop.html', {'form': form})


def register_workshop(request, workshop_id):
    workshop = allWorkshop.objects.get(id=workshop_id)

    if request.method == 'POST':
        form = WorkshopRegistrationForm(request.POST)
        if form.is_valid():
            registration = form.save(commit=False)
            registration.workshop = workshop
            registration.save()
            # redirect to razorpay
            messages.success(request, 'Enrolled successfully!')
    else:
        form = WorkshopRegistrationForm()

    context = {
        'workshop': workshop,
        'form': form
    }
    return render(request, 'registration_form.html', context)


def workshop_detail(request, workshop_id):
    workshop = allWorkshop.objects.get(id=workshop_id)
    return render(request, 'workshop_detail.html', {'workshop': workshop})


def category_detail_view(request, category_id):
    category = workshopCategory.objects.get(id=category_id)
    return render(request, 'category_detail.html', {'category': category})


def search_workshops(request):
    query = request.GET.get('query')
    workshops = allWorkshop.objects.filter(title__icontains=query)
    return render(request, 'search_results.html', {'workshops': workshops, 'query': query})


def submit_review(request):
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Review submitted successfully!')
            form = ReviewForm()  # Reset the form after submission
    else:
        form = ReviewForm()
    return render(request, 'add_review.html', {'form': form})


def blog_list(request):
    blogs = Blog.objects.filter(is_visible=False)
    return render(request, 'blog_view.html', {'blogs': blogs})

